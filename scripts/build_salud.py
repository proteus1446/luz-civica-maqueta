#!/usr/bin/env python3
"""
Genera data/data_salud.js a partir de datos reales SINIM (4-SALUD.xlsx), para
las 345 comunas de Chile, años 2008-2025.

A diferencia de Administración/Dotación, esta página consume los códigos SINIM
"en crudo" (sin renombrar a un esquema propio) — el JS de maqueta_salud.html
lee directamente d.ISAL019, d.MPSP, etc. Este script simplemente vuelca las
columnas usadas por la página, ya identificadas comparando el DATA embebido
original (5 comunas de muestra) contra las columnas reales del archivo SINIM.

Nota: MVACU y MTFCOTR existen en el archivo SINIM y en la especificación de
variables, pero el JS de la página NO los referencia (grep confirmó 0 usos) —
se omiten a propósito para no incluir datos que la página no consume.

MASM (S-N) ¿Administra Servicio de Salud Primaria? se agrega para poder
mostrar un aviso claro ("esta comuna no administra Salud") en vez de
tarjetas llenas de "Sin dato" cuando MASM es "No" / "Sin Servicio" / "No
Recepcionado" (a diferencia de MTAS, que describe QUIÉN administra y no
sirve como bandera confiable: trae "Sin Servicio" en algunos casos donde
igual hay datos numéricos).

HPISM en 0 se guarda como null: ninguna comuna real tiene 0 personas
inscritas en FONASA, es la misma convención de dato faltante que usa SINIM
para esta columna.

ISAL005 (Cobertura de Salud Primaria Municipal) = HPISM/ITPC*100. Cuando
SINIM trae HPISM pero no ISAL005, se completa con esta fórmula usando la
población (ITPC de 4-SALUD.xlsx no existe como columna — se usa ICAR004 de
7-caracterizacion comunal.xlsx, que es la misma cifra exacta que ITPC en
las 6.210 filas donde ambas están disponibles).
"""
import json
import re
import openpyxl

from build_administracion import comuna_key, num

SINIM_SALUD = "/Users/cristobal/prueba/sinim/4-SALUD.xlsx"
SINIM_CARAC = "/Users/cristobal/prueba/sinim/7-caracterizacion comunal.xlsx"
OUT_JS = "/Users/cristobal/Downloads/maqueta/data/data_salud.js"


def load_poblacion():
    """ICAR004 (7-caracterizacion comunal.xlsx) — idéntico a ITPC de la misma
    planilla (verificado exacto en las 6.210 filas comuna-año), que es la
    población que usa la fórmula real de ISAL005: HPISM/ITPC*100."""
    wb = openpyxl.load_workbook(SINIM_CARAC, read_only=True)
    ws = wb["Hoja1"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx_pob = [i for i, h in enumerate(header) if h and str(h).startswith("ICAR004")][0]
    idx_anio = len(header) - 1
    out = {}
    for r in rows:
        if r[0] is None:
            continue
        municipio = comuna_key(str(r[1]).strip())
        anio = str(r[idx_anio])
        out[(municipio, anio)] = num(r[idx_pob])
    return out

# Códigos numéricos usados por maqueta_salud.html (confirmado por inspección del
# DATA embebido original + grep del JS). MTAS se maneja aparte por ser texto.
CODES_NUM = [
    "GTCM", "HPISM", "ISAL005", "ISAL009", "ISAL010", "ISAL012", "ISAL013",
    "ISAL015", "ISAL018", "ISAL019", "ISAL021", "ISAL023", "ISAL025",
    "ISAL029", "ISAL031", "ISAL032", "ISAL23", "MAMBUL", "MCECOF", "MCESFAM",
    "MCOSAM", "MDENTAL", "MNCGR", "MNCGU", "MNPR", "MPSCC", "MPSCDT", "MPSH",
    "MPSOC", "MPSP", "MSAPU", "MSFARM", "MSOPT", "MTFCE", "MTFCM", "MTFFOND",
    "MTFGER", "MTFKINE", "MTFMATRO", "MTFNUTRI", "MTFODON", "MTFPSICO",
    "MTFPSIQ", "MTFTECENF", "MTFTECMED",
]
CODE_TEXT = "MTAS"
# MASM (S-N): ¿administra Servicio de Salud Primaria? "Si" / "No" / "Sin
# Servicio" / "No Recepcionado". Se usa para mostrar un aviso claro en vez de
# tarjetas llenas de "Sin dato" cuando la comuna no administra Salud.
CODE_TEXT_MASM = "MASM"


def find_col(header, code):
    pat = r"(?<![A-Za-z0-9])" + re.escape(code) + r"(?!\.\d)(?![A-Za-z0-9])"
    for i, h in enumerate(header):
        if h and re.search(pat, str(h)):
            return i
    raise KeyError(f"Column not found for code {code}")


def find_anio_col(header):
    for i, h in enumerate(header):
        if h and str(h).strip().upper() == "AÑO":
            return i
    raise KeyError("Año column not found")


def main():
    wb = openpyxl.load_workbook(SINIM_SALUD, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)

    idx_num = {c: find_col(header, c) for c in CODES_NUM}
    idx_text = find_col(header, CODE_TEXT)
    idx_masm = find_col(header, CODE_TEXT_MASM)
    idx_anio = find_anio_col(header)
    poblacion = load_poblacion()

    data = {}
    isal005_completados = 0
    for r in rows:
        if r[0] is None:
            continue
        municipio = comuna_key(str(r[1]).strip())
        anio = str(r[idx_anio])
        record = {c: num(r[i]) for c, i in idx_num.items()}
        # HPISM (población inscrita FONASA) en 0 no es un valor real — ninguna
        # comuna tiene 0 inscritos; es la misma convención de "sin dato" que
        # usa el resto del archivo SINIM para esta columna en particular.
        if record.get("HPISM") == 0:
            record["HPISM"] = None
        # ISAL005 (Cobertura de Salud Primaria Municipal) = HPISM/ITPC*100.
        # Cuando SINIM no trae ISAL005 pero sí HPISM, se completa con la
        # fórmula real usando la población (ITPC/ICAR004, misma cifra —
        # verificado exacto en las 6.210 filas comuna-año donde ambas
        # columnas están presentes: diferencia media de 0,002 puntos,
        # solo redondeo).
        if record.get("ISAL005") is None and record.get("HPISM") is not None:
            pob = poblacion.get((municipio, anio))
            if pob:
                record["ISAL005"] = round(record["HPISM"] / pob * 100, 2)
                isal005_completados += 1
        raw_text = r[idx_text]
        record[CODE_TEXT] = raw_text if isinstance(raw_text, str) else None
        raw_masm = r[idx_masm]
        record[CODE_TEXT_MASM] = raw_masm if isinstance(raw_masm, str) else None
        data.setdefault(municipio, {})[anio] = record

    with open(OUT_JS, "w", encoding="utf-8") as f:
        f.write("// Generado por scripts/build_salud.py — no editar a mano.\n")
        f.write("const DATA_SALUD = ")
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")

    print(f"ISAL005 completados con HPISM/población: {isal005_completados}")

    n_comunas = len(data)
    n_rows = sum(len(v) for v in data.values())
    print(f"OK: {n_comunas} comunas, {n_rows} filas comuna-año -> {OUT_JS}")


if __name__ == "__main__":
    main()
