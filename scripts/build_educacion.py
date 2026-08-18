#!/usr/bin/env python3
"""
Genera data/data_educacion.js a partir de datos reales SINIM, para las 345
comunas de Chile, años 2008-2025.

Fuentes:
  - /Users/cristobal/prueba/sinim/3-Educacion.xlsx                 (todo el detalle del sector)
  - /Users/cristobal/prueba/sinim/1-Administracion_finanzas.xlsx   (ingreso_municipal_total = IADM999)
  - /Users/cristobal/prueba/sinim/7-caracterizacion comunal.xlsx   (población, ICAR004)

Fórmulas (verificadas exactas contra Providencia 2008/2009 hardcodeado):
  activa                  = bool(IEDU025)  (mismo criterio que areas_activas.educacion en Dotación)
  edad_escolar             = IPEEC
  cobertura_pct            = IEDU009
  asistencia_pct           = IEDU005
  matricula                = DNAM
  establecimientos         = IEDU002
  docentes_aula             = DNDA
  alumnos_por_docente      = DNAM / DNDA
  gasto_alumno_anual       = IEDU025 / DNAM
  gasto_alumno_mensual     = IEDU025 / DNAM / 12
  dependencia_subvencion_pct = IEDU019
  ingresos.subvencion      = IEDU018
  ingresos.aporte_municipal = IEDU020
  ingresos.total           = IEDU999
  ingresos.otros           = total - subvencion - aporte_municipal   (residual)
  gastos.personal          = IEDU026
  gastos.operacional       = IEDU029
  gastos.inversion         = IEDU031
  gastos.total             = IEDU025
  gastos.otros             = total - personal - operacional - inversion   (residual)
  personal_funcion.docentes    = MTPD
  personal_funcion.no_docentes = MTPND
  personal_funcion.total       = docentes + no_docentes
  personal_contrato.planta     = IEDU040
  personal_contrato.contrata   = IEDU042
  personal_contrato.cdt        = IEDU041
  personal_contrato.honorarios = IEDU043
  personal_contrato.total      = suma de las 4 categorías
  ingreso_municipal_total       = IADM999 (archivo de Administración)
  aporte_municipal_pct_gasto_muni = None si aporte_municipal es 0/None, si no
    round(aporte_municipal / ingreso_municipal_total * 100, 2)
    (el nombre del campo dice "gasto_muni" pero numéricamente usa el ingreso
    municipal total, no el gasto — confirmado reproduciendo 3.78% exacto)
"""
import json
import re
import openpyxl

from build_administracion import comuna_key, num

SINIM_EDU = "/Users/cristobal/prueba/sinim/3-Educacion.xlsx"
SINIM_ADMIN = "/Users/cristobal/prueba/sinim/1-Administracion_finanzas.xlsx"
SINIM_CARAC = "/Users/cristobal/prueba/sinim/7-caracterizacion comunal.xlsx"
OUT_JS = "/Users/cristobal/Downloads/maqueta/data/data_educacion.js"

CODES = ["DNAM", "DNDA", "IEDU002", "IEDU005", "IEDU009", "IEDU018", "IEDU019",
         "IEDU020", "IEDU025", "IEDU026", "IEDU029", "IEDU031", "IEDU999",
         "IPEEC", "MTPD", "MTPND", "IEDU040", "IEDU041", "IEDU042", "IEDU043"]


def find_col(header, code):
    pat = r"(?<![A-Za-z0-9])" + re.escape(code) + r"(?!\.\d)(?![A-Za-z0-9])"
    for i, h in enumerate(header):
        if h and re.search(pat, str(h)):
            return i
    raise KeyError(f"Column not found for code {code}")


def find_anio_col(header):
    for i, h in enumerate(header):
        if h and str(h).strip().upper() == "AÑO":
            return i
    raise KeyError("Año column not found")


def load_sheet(path, codes):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {c: find_col(header, c) for c in codes}
    idx_anio = find_anio_col(header)
    out = {}
    for r in rows:
        if r[0] is None:
            continue
        municipio = comuna_key(str(r[1]).strip())
        anio = str(r[idx_anio])
        out[(municipio, anio)] = {c: num(r[i]) for c, i in idx.items()}
    return out


def load_poblacion():
    wb = openpyxl.load_workbook(SINIM_CARAC, read_only=True)
    ws = wb["Hoja1"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx_pob = [i for i, h in enumerate(header) if h and str(h).startswith("ICAR004")][0]
    idx_anio = len(header) - 1
    out = {}
    for r in rows:
        if r[0] is None:
            continue
        municipio = comuna_key(str(r[1]).strip())
        anio = str(r[idx_anio])
        out[(municipio, anio)] = num(r[idx_pob])
    return out


def main():
    edu = load_sheet(SINIM_EDU, CODES)
    admin = load_sheet(SINIM_ADMIN, ["IADM999"])
    poblacion = load_poblacion()

    data = {}
    for (municipio, anio), e in edu.items():
        def g(k):
            return e.get(k) or 0

        matricula = e.get("DNAM")
        docentes_aula = e.get("DNDA")
        gasto_total = e.get("IEDU025")

        alumnos_por_docente = round(matricula / docentes_aula, 3) if matricula and docentes_aula else None
        gasto_alumno_anual = round(gasto_total / matricula, 3) if gasto_total is not None and matricula else None
        gasto_alumno_mensual = round(gasto_total / matricula / 12, 3) if gasto_total is not None and matricula else None

        ing_total = e.get("IEDU999")
        ing_subv = e.get("IEDU018")
        ing_muni = e.get("IEDU020")
        ing_otros = (ing_total - g("IEDU018") - g("IEDU020")) if ing_total is not None else None

        gas_personal = e.get("IEDU026")
        gas_operacional = e.get("IEDU029")
        gas_inversion = e.get("IEDU031")
        gas_otros = (gasto_total - g("IEDU026") - g("IEDU029") - g("IEDU031")) if gasto_total is not None else None

        docentes = e.get("MTPD")
        no_docentes = e.get("MTPND")

        planta = g("IEDU040")
        contrata = g("IEDU042")
        cdt = g("IEDU041")
        honorarios = g("IEDU043")

        ingreso_municipal_total = admin.get((municipio, anio), {}).get("IADM999")
        aporte_pct = (round(ing_muni / ingreso_municipal_total * 100, 2)
                      if ing_muni and ingreso_municipal_total else None)

        data.setdefault(municipio, {})[anio] = {
            "poblacion": poblacion.get((municipio, anio)),
            "activa": bool(gasto_total),
            "edad_escolar": e.get("IPEEC"),
            "cobertura_pct": e.get("IEDU009"),
            "asistencia_pct": e.get("IEDU005"),
            "matricula": matricula,
            "establecimientos": e.get("IEDU002"),
            "docentes_aula": docentes_aula,
            "alumnos_por_docente": alumnos_por_docente,
            "gasto_alumno_anual": gasto_alumno_anual,
            "gasto_alumno_mensual": gasto_alumno_mensual,
            "dependencia_subvencion_pct": e.get("IEDU019"),
            "ingresos": {
                "subvencion": ing_subv,
                "aporte_municipal": ing_muni,
                "otros": ing_otros,
                "total": ing_total,
            },
            "gastos": {
                "personal": gas_personal,
                "operacional": gas_operacional,
                "inversion": gas_inversion,
                "otros": gas_otros,
                "total": gasto_total,
            },
            "personal_funcion": {
                "docentes": docentes,
                "no_docentes": no_docentes,
                "total": (docentes or 0) + (no_docentes or 0) if (docentes is not None or no_docentes is not None) else None,
            },
            "personal_contrato": {
                "planta": planta,
                "contrata": contrata,
                "cdt": cdt,
                "honorarios": honorarios,
                "total": planta + contrata + cdt + honorarios,
            },
            "ingreso_municipal_total": ingreso_municipal_total,
            "aporte_municipal_pct_gasto_muni": aporte_pct,
        }

    with open(OUT_JS, "w", encoding="utf-8") as f:
        f.write("// Generado por scripts/build_educacion.py — no editar a mano.\n")
        f.write("const DATA_EDUCACION = ")
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")

    n_comunas = len(data)
    n_rows = sum(len(v) for v in data.values())
    print(f"OK: {n_comunas} comunas, {n_rows} filas comuna-año -> {OUT_JS}")


if __name__ == "__main__":
    main()
