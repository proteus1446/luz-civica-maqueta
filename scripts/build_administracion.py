#!/usr/bin/env python3
"""
Genera data/data_administracion.js a partir de datos reales SINIM + Contraloría,
para las 345 comunas de Chile, años 2008-2025.

Fuentes:
  - /Users/cristobal/prueba/sinim/1-Administracion_finanzas.xlsx  (ingresos/gastos/kpis SINIM)
  - /Users/cristobal/prueba/sinim/7-caracterizacion comunal.xlsx  (población, ICAR004)
  - /Users/cristobal/Desktop/luz _c/contraloria .xlsx             (deficit, situacion, deuda flotante)

Fórmulas (verificadas contra los valores hardcodeados originales de Providencia 2008):
  poblacion              = ICAR004
  kpis.dependencia_fcm   = IADM75  (%)
  kpis.ejecucion         = IADM 125 (%)
  kpis.eficiencia_cobro  = IADM100 (%)
  kpis.deuda_flotante_pct= pagado/flotante*100 (100 si flotante es 0/nulo, null si no hay fila)
  ingresos.ipp           = IADM41
  ingresos.traspaso_fcm  = "Ingreso para fondo comun" (Contraloría) — NO
    usa IADM39 como respaldo (IADM39 es una cuenta de GASTO, no de
    ingreso; solo corresponde para gastos.fcm más abajo). Si la comuna/año
    no está en el archivo de Contraloría, se asume 0 — no "sin dato" — y
    Otros ingresos absorbe todo el residuo. Verificado exacto contra el
    caso de prueba (Providencia 2024: $65.818.459,544). Con tope al
    residuo disponible en el 0,2% de filas donde excede el residuo grueso
    — ver ingresos.otros
  ingresos.transferencias= IADM43.1 (solo desde 2011)
  ingresos.fcm_recibido  = IADM40
  ingresos.total         = IADM999
  ingresos.otros         = (total - ipp - fcm_recibido - transferencias) - traspaso_fcm
    (residual, restando TAMBIÉN traspaso_fcm — antes NO se restaba y el
    monto quedaba contado dos veces: una vez en su propia línea y otra
    escondido dentro de "Otros ingresos"; ver bug_doble_conteo_ingresos.md
    para el caso de prueba completo. Nunca queda negativo: en el 0,2% de
    filas donde traspaso_fcm > residuo grueso, por descalce entre fuentes
    (Contraloría vs. SINIM del resto de las cuentas), se topa el traspaso
    mostrado en Ingresos al residuo disponible y Otros queda en 0 —
    gastos.fcm más abajo usa IADM39 tal cual, no le aplica este ajuste)
  gastos.fcm             = IADM39
  gastos.personal        = IADM61
  gastos.bienes_servicios= IADM85
  gastos.salud           = IADM77
  gastos.inversion       = IADM22
  gastos.educacion       = IADM76
  gastos.total           = IADM11
  gastos.otros           = total - fcm - personal - bienes_servicios - salud - inversion - educacion (residual)
  deficit, situacion, deuda_flotante, deuda_flotante_pagado -> Contraloría
"""
import json
import unicodedata
import openpyxl


# Variantes de nombre entre el archivo SINIM y el de Contraloría para la misma
# comuna (verificado manualmente: las 4 quedaban sin región/deficit/deuda flotante
# por el desajuste de nombre, no por falta real de dato).
ALIASES_CONTRALORIA = {
    "PAIHUANO": "PAIGUANO",
    "LA CALERA": "CALERA",
    "LLAY LLAY": "LLAILLAY",
    "O'HIGGINS": "O´HIGGINS",
}


def comuna_key(municipio):
    """Normalize a comuna name to the site's canonical key: uppercase, vowel
    accents stripped, but Ñ preserved (matches MAIPU/ÑUÑOA/PROVIDENCIA convention
    already used across panel_comunal.html and friends)."""
    s = str(municipio).strip().upper()
    s = s.replace("Ñ", "")  # placeholder so NFD doesn't decompose it
    s = unicodedata.normalize("NFD", s)
    s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
    s = s.replace("", "Ñ")
    return s

SINIM_ADMIN = "/Users/cristobal/prueba/sinim/1-Administracion_finanzas.xlsx"
SINIM_CARAC = "/Users/cristobal/prueba/sinim/7-caracterizacion comunal.xlsx"
CONTRALORIA = "/Users/cristobal/Desktop/luz _c/contraloria .xlsx"
OUT_JS = "/Users/cristobal/Downloads/maqueta/data/data_administracion.js"
OUT_NOMBRES = "/Users/cristobal/Downloads/maqueta/data/nombres_comunas.js"
OUT_REGIONES = "/Users/cristobal/Downloads/maqueta/data/regiones_comunas.js"


def num(v):
    """Coerce a raw cell value to float/int or None. SINIM exports use strings like
    'No Recepcionado' or '.' for missing data; treat anything non-numeric as None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if not s or s in (".", "-", "No Recepcionado", "S/I", "S/R"):
        return None
    s = s.replace(".", "").replace(",", ".") if s.count(".") > 1 else s
    try:
        return float(s)
    except ValueError:
        return None


def compute_otros_ing(total_ing, ipp, fcm_recibido, transferencias, fcm_traspaso):
    """"Otros ingresos" como residuo, restando también el traspaso al FCM
    (si no, ese monto queda contado dos veces — ver
    bug_doble_conteo_ingresos.md). Devuelve (fcm_traspaso_mostrado,
    otros_ing); el primero viene topado al residuo disponible en el caso
    raro de descalce entre fuentes, para no mostrar Otros negativo."""
    if total_ing is None:
        return fcm_traspaso, None
    residuo_grueso = total_ing - (ipp or 0) - (fcm_recibido or 0) - (transferencias or 0)
    if fcm_traspaso is None:
        return fcm_traspaso, residuo_grueso
    if fcm_traspaso <= residuo_grueso:
        return fcm_traspaso, residuo_grueso - fcm_traspaso
    return residuo_grueso, 0.0


def col_index(header, code):
    """Find the column index whose header starts with the exact SINIM code."""
    code_norm = code.strip()
    for i, h in enumerate(header):
        if not h:
            continue
        h = str(h)
        # header format: "IADM41 (M$) Nombre..." -> code is first token
        token = h.split(" ", 1)[0]
        if token == code_norm:
            return i
    raise KeyError(f"Column not found for code {code}")


def load_admin():
    wb = openpyxl.load_workbook(SINIM_ADMIN, read_only=True)
    ws = wb["2025-2008"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)

    idx = {
        "codigo": 0,
        "municipio": 1,
        "ejecucion": col_index(header, "IADM"),  # placeholder, fixed below
    }
    # exact lookups
    c_ejecucion = col_index(header, "IADM")
    # Because "IADM 125" has a space, split token won't match "IADM" alone; do explicit search.
    def find(code):
        for i, h in enumerate(header):
            if h and str(h).replace(" ", "").startswith(code.replace(" ", "") + "("):
                return i
        raise KeyError(code)

    c = {
        "ejecucion": find("IADM 125"),
        "eficiencia_cobro": find("IADM100"),
        "ipp": find("IADM41"),
        "traspaso_fcm": find("IADM39"),
        "transferencias": find("IADM43.1"),
        "fcm_recibido": find("IADM40"),
        "total_ing": find("IADM999"),
        "personal": find("IADM61"),
        "bienes_servicios": find("IADM85"),
        "salud": find("IADM77"),
        "inversion": find("IADM22"),
        "educacion": find("IADM76"),
        "total_gas": find("IADM11"),
        "dependencia_fcm": find("IADM75"),
        "anio": len(header) - 1,  # 'Año' is last column
    }

    data = {}
    nombres = {}
    for r in rows:
        if r[0] is None:
            continue
        codigo = r[0]
        municipio_raw = str(r[1]).strip()
        municipio = comuna_key(municipio_raw)
        anio = str(r[c["anio"]])
        nombres[municipio] = municipio_raw.title().replace(" De ", " de ").replace(" Del ", " del ")

        def g(key):
            return num(r[c[key]])

        total_ing = g("total_ing")
        ipp = g("ipp")
        fcm_recibido = g("fcm_recibido")
        transferencias = g("transferencias")
        fcm_g = g("traspaso_fcm")  # IADM39 — cuenta de GASTO, usada solo en gastos.fcm más abajo
        # ingresos.traspaso_fcm NO usa IADM39 (es una cuenta de gasto, no de
        # ingreso — no corresponde como respaldo acá). La fuente real es
        # "Ingreso para fondo comun" (Contraloría), que add_contraloria()
        # aplica más abajo. Si esa comuna/año no está en Contraloría, se
        # asume 0 (no "sin dato"): Otros ingresos absorbe todo el residuo.
        fcm_ing_mostrado, otros_ing = compute_otros_ing(total_ing, ipp, fcm_recibido, transferencias, 0)

        total_gas = g("total_gas")
        personal = g("personal")
        bienes = g("bienes_servicios")
        salud = g("salud")
        inversion = g("inversion")
        educacion = g("educacion")
        otros_gas = None
        if total_gas is not None:
            otros_gas = total_gas - (fcm_g or 0) - (personal or 0) - (bienes or 0) - (salud or 0) - (inversion or 0) - (educacion or 0)

        data.setdefault(municipio, {})[anio] = {
            "poblacion": None,  # filled later
            "kpis": {
                "dependencia_fcm": g("dependencia_fcm"),
                "ejecucion": g("ejecucion"),
                "deuda_flotante_pct": None,  # filled later
                "eficiencia_cobro": g("eficiencia_cobro"),
            },
            "ingresos": {
                "ipp": ipp,
                "traspaso_fcm": fcm_ing_mostrado,
                "transferencias": transferencias,
                "fcm_recibido": fcm_recibido,
                "otros": otros_ing,
                "total": total_ing,
            },
            "gastos": {
                "fcm": fcm_g,
                "personal": personal,
                "bienes_servicios": bienes,
                "salud": salud,
                "inversion": inversion,
                "educacion": educacion,
                "otros": otros_gas,
                "total": total_gas,
            },
            "deficit": None,       # filled later
            "situacion": None,     # filled later
            "deuda_flotante": None,        # filled later
            "deuda_flotante_pagado": None, # filled later
            "_codigo": codigo,
        }
    return data, nombres


def add_poblacion(data):
    wb = openpyxl.load_workbook(SINIM_CARAC, read_only=True)
    ws = wb["Hoja1"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx_pob = [i for i, h in enumerate(header) if h and str(h).startswith("ICAR004")][0]
    idx_anio = len(header) - 1
    for r in rows:
        if r[0] is None:
            continue
        municipio = comuna_key(str(r[1]).strip())
        anio = str(r[idx_anio])
        if municipio in data and anio in data[municipio]:
            data[municipio][anio]["poblacion"] = num(r[idx_pob])


REGIONES = {
    "XV": "Arica y Parinacota", "I": "Tarapacá", "II": "Antofagasta",
    "III": "Atacama", "IV": "Coquimbo", "V": "Valparaíso", "RM": "Metropolitana",
    "VI": "O'Higgins", "VII": "Maule", "XVI": "Ñuble", "VIII": "Biobío",
    "IX": "Araucanía", "XIV": "Los Ríos", "X": "Los Lagos", "XI": "Aysén",
    "XII": "Magallanes",
}
# orden geográfico norte -> sur, el mismo que usa Contraloría/SUBDERE
ORDEN_REGIONES = ["XV", "I", "II", "III", "IV", "V", "RM", "VI", "VII", "XVI",
                   "VIII", "IX", "XIV", "X", "XI", "XII"]


def add_contraloria(data):
    wb = openpyxl.load_workbook(CONTRALORIA, read_only=True)
    ws = wb["Sheet1"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {h: i for i, h in enumerate(header)}
    regiones_comuna = {}
    for r in rows:
        if r[0] is None:
            continue
        anio = str(r[idx["Año"]])
        municipio = comuna_key(str(r[idx["Nombre Municipio"]]).strip())
        municipio = ALIASES_CONTRALORIA.get(municipio, municipio)
        deficit = num(r[idx["Deficit/super"]])
        situacion = r[idx["Situacion"]]
        flotante = num(r[idx["Deuda Flotante"]])
        pagado = num(r[idx["Deuda Flotante pagado"]])
        region = str(r[idx["Región"]]).strip()
        if region == "XIII":
            region = "RM"

        if flotante is None:
            pct = None
        elif flotante == 0:
            pct = 100.0
        else:
            pct = round((pagado or 0) / flotante * 100, 2)

        fondo_comun = num(r[idx["Ingreso para fondo comun"]])

        if municipio in data and anio in data[municipio]:
            d = data[municipio][anio]
            d["deficit"] = deficit
            d["situacion"] = situacion
            d["deuda_flotante"] = flotante
            d["deuda_flotante_pagado"] = pagado
            d["kpis"]["deuda_flotante_pct"] = pct
            # "Ingreso para fondo comun" (Contraloría) es la fuente de
            # ingresos.traspaso_fcm — reemplaza el 0 asumido por defecto en
            # load_admin() (para comunas/años sin fila en Contraloría) y
            # recalcula ingresos.otros con el mismo criterio (residuo, sin
            # negativos). gastos.fcm no se toca acá, sigue en IADM39.
            if fondo_comun is not None:
                ing = d["ingresos"]
                ing["traspaso_fcm"], ing["otros"] = compute_otros_ing(
                    ing["total"], ing["ipp"], ing["fcm_recibido"], ing["transferencias"], fondo_comun)

        # nos quedamos con la región del año más reciente vista para esa comuna
        if region in REGIONES:
            prev_year = regiones_comuna.get(municipio, (None, None))[1]
            if prev_year is None or anio >= prev_year:
                regiones_comuna[municipio] = (region, anio)

    return {k: v[0] for k, v in regiones_comuna.items()}


def main():
    data, nombres = load_admin()
    add_poblacion(data)
    region_por_comuna = add_contraloria(data)

    # strip internal helper key
    for muni in data:
        for anio in data[muni]:
            data[muni][anio].pop("_codigo", None)

    with open(OUT_JS, "w", encoding="utf-8") as f:
        f.write("// Generado por scripts/build_administracion.py — no editar a mano.\n")
        f.write("const DATA_ADMINISTRACION = ")
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")

    with open(OUT_NOMBRES, "w", encoding="utf-8") as f:
        f.write("// Generado por scripts/build_administracion.py — no editar a mano.\n")
        f.write("const NOMBRES_COMUNAS = ")
        json.dump(nombres, f, ensure_ascii=False, indent=2, sort_keys=True)
        f.write(";\n")

    with open(OUT_REGIONES, "w", encoding="utf-8") as f:
        f.write("// Generado por scripts/build_administracion.py — no editar a mano.\n")
        f.write("const NOMBRES_REGIONES = ")
        json.dump(REGIONES, f, ensure_ascii=False, indent=2)
        f.write(";\n")
        f.write("const ORDEN_REGIONES = ")
        json.dump(ORDEN_REGIONES, f, ensure_ascii=False)
        f.write(";\n")
        f.write("const REGION_POR_COMUNA = ")
        json.dump(region_por_comuna, f, ensure_ascii=False, indent=2, sort_keys=True)
        f.write(";\n")

    n_sin_region = len([k for k in data if k not in region_por_comuna])
    if n_sin_region:
        print(f"AVISO: {n_sin_region} comunas sin región asignada")

    n_comunas = len(data)
    n_rows = sum(len(v) for v in data.values())
    print(f"OK: {n_comunas} comunas, {n_rows} filas comuna-año -> {OUT_JS}")


if __name__ == "__main__":
    main()
