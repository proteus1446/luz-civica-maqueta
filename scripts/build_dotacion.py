#!/usr/bin/env python3
"""
Genera data/data_dotacion.js a partir de datos reales SINIM, para las 345 comunas
de Chile, años 2008-2025.

Fuentes:
  - /Users/cristobal/prueba/sinim/1-Administracion_finanzas.xlsx  (gasto personal municipal, límites legales)
  - /Users/cristobal/prueba/sinim/2-Recursos H.xlsx                (headcount municipal, % profesionalización/femenina)
  - /Users/cristobal/prueba/sinim/3-Educacion.xlsx                 (headcount y gasto personal Educación)
  - /Users/cristobal/prueba/sinim/4-SALUD.xlsx                     (headcount y gasto personal Salud)
  - /Users/cristobal/prueba/sinim/7-caracterizacion comunal.xlsx   (población, ICAR004)

Fórmulas (verificadas al número exacto contra los valores hardcodeados originales
de Providencia 2008):
  poblacion                     = ICAR004
  profesionalizacion_pct        = IADM25, normalizado a escala 0-100 (SINIM
    lo entrega en fracción 0-1 hasta 2023 y en 0-100 desde 2024 — se
    detecta y corrige automáticamente, ver pct_norm())
  participacion_femenina_pct    = IADM33 (mismo tratamiento que IADM25)
  municipal_total   = IRH05 + IRH12 + IRH15 + IRH16               (planta+contrata+honorarios+comunitarios; IRH16 solo existe desde 2019)
  educacion_total   = IEDU040 + IEDU042 + IEDU043 + IEDU041       (+ cdt)
  salud_total       = MPSP + MPSCC + MPSH + MPSCDT + MPSOC        (+ cdt + comunitarios)
  consolidado_total = municipal_total + educacion_total + salud_total
  consolidado.planta       = IRH05 + IEDU040 + MPSP
  consolidado.contrata     = IRH12 + IEDU042 + MPSCC
  consolidado.cdt          = IEDU041 + MPSCDT
  consolidado.honorarios   = IRH15 + IEDU043 + MPSH
  consolidado.comunitarios = IRH16 + MPSOC
  gasto.planta       = IADM78 + IEDU040.1 + ISAL029
  gasto.contrata      = IADM79 + IEDU042.1 + ISAL031
  gasto.honorarios    = IADM80 + IEDU043.1 + ISAL032
  gasto.comunitarios  = IADM111  (solo categoría municipal)
  gasto.total         = IADM61 + IEDU026.1 + ISAL019.1
  limites.lim42 (gasto en personal)   = (IADM61 / (IADM42 * 0.42)) * 100
    "Participación de Gastos en Personal Respecto del Umbral Legal (42%)",
    criterio SUBDERE. Fuente: Balance de Ejecución Presupuestario (BEP).
    Base legal: Art. 1 Ley 18.294, modificado por la letra a) del Art. 65
    de la Ley 18.382 — el gasto anual en personal no puede exceder el 42%
    del rendimiento estimado de los ingresos propios municipales. Incluye
    la renta del alcalde/alcaldesa (va dentro del Subtítulo 21 — IADM61 —
    y no se puede aislar). Sobre 100% excede el umbral del 42%.
  limites.lim40 (personal a contrata) = (IADM79 / (IADM78 * 0.40)) * 100
  limites.lim10 (honorarios)          = (IADM80 / IADM78) * 100 — se compara
    contra 10 (no contra 100). Fórmula del Art. 13, Ley N° 19.280.
    (Hasta [fecha del fix de esta línea] se usaba (IADM80/(IADM61*0.10))*100,
    que por casualidad reproducía el valor hardcodeado de Providencia 2008
    (86,43) pero NO es la fórmula legal — daba resultados absurdos como
    46,46% "dentro del límite" para un caso que en realidad excede el 10%
    real por poco. Se corrige a la fórmula correcta de la ley aunque ya no
    reproduzca ese valor histórico.)
  limites.blindspot_comunitarios_pct = ((IADM80 + IADM111) / IADM78) * 100
    (solo Personal Municipal; NO es un límite legal, no tiene techo — ver
    nota junto al cálculo, más abajo en el código)
  municipal_honorarios_sindato = True si IRH15 falta/0 pero IADM80 (gasto) > 0
  areas_activas.educacion/salud = True si el gasto en personal del sector es
    > 0 O si hay headcount del sector > 0 (28 filas de educación y 25 de
    salud tienen headcount real con gasto en 0 — antes solo se miraba el
    gasto y esas personas "desaparecían" de la tarjeta de Dotación)
  gasto_por_area.{municipal,educacion,salud} = IADM61, IEDU026.1, ISAL019.1
    (campo usado por el gráfico de evolución "Municipal/Educación/Salud" del HTML;
    no existía en los datos originales — el gráfico estaba roto desde antes)
"""
import json
import re
import openpyxl

from build_administracion import comuna_key, num  # reutiliza normalización ya validada

SINIM_ADMIN = "/Users/cristobal/prueba/sinim/1-Administracion_finanzas.xlsx"
SINIM_RRHH = "/Users/cristobal/prueba/sinim/2-Recursos H.xlsx"
SINIM_EDU = "/Users/cristobal/prueba/sinim/3-Educacion.xlsx"
SINIM_SALUD = "/Users/cristobal/prueba/sinim/4-SALUD.xlsx"
SINIM_CARAC = "/Users/cristobal/prueba/sinim/7-caracterizacion comunal.xlsx"
OUT_JS = "/Users/cristobal/Downloads/maqueta/data/data_dotacion.js"


def find_col(header, code):
    """Ubica la columna cuyo header contiene el código SINIM como token exacto,
    en cualquier posición (algunos archivos lo ponen al inicio, otros al final)."""
    pat = r"(?<![A-Za-z0-9])" + re.escape(code) + r"(?!\.\d)(?![A-Za-z0-9])"
    for i, h in enumerate(header):
        if h and re.search(pat, str(h)):
            return i
    raise KeyError(f"Column not found for code {code}")


def find_anio_col(header):
    for i, h in enumerate(header):
        if h and str(h).strip().upper() == "AÑO":
            return i
    raise KeyError("Año column not found")


def load_sheet(path, codes):
    """Devuelve {(comuna_key, anio): {code: valor_numerico}} para los códigos pedidos."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {c: find_col(header, c) for c in codes}
    idx_anio = find_anio_col(header)
    out = {}
    for r in rows:
        if r[0] is None:
            continue
        municipio = comuna_key(str(r[1]).strip())
        anio = str(r[idx_anio])
        out[(municipio, anio)] = {c: num(r[i]) for c, i in idx.items()}
    return out


def load_poblacion():
    wb = openpyxl.load_workbook(SINIM_CARAC, read_only=True)
    ws = wb["Hoja1"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx_pob = [i for i, h in enumerate(header) if h and str(h).startswith("ICAR004")][0]
    idx_anio = len(header) - 1
    out = {}
    for r in rows:
        if r[0] is None:
            continue
        municipio = comuna_key(str(r[1]).strip())
        anio = str(r[idx_anio])
        out[(municipio, anio)] = num(r[idx_pob])
    return out


def main():
    admin = load_sheet(SINIM_ADMIN, ["IADM78", "IADM79", "IADM80", "IADM111", "IADM61", "IADM42"])
    rrhh = load_sheet(SINIM_RRHH, ["IRH05", "IRH12", "IRH15", "IRH16", "IADM25", "IADM33"])
    edu = load_sheet(SINIM_EDU, ["IEDU040", "IEDU042", "IEDU043", "IEDU041",
                                  "IEDU040.1", "IEDU042.1", "IEDU043.1", "IEDU026.1"])
    salud = load_sheet(SINIM_SALUD, ["MPSP", "MPSCC", "MPSH", "MPSCDT", "MPSOC",
                                      "ISAL029", "ISAL031", "ISAL032", "ISAL019.1"])
    poblacion = load_poblacion()

    keys = sorted(set(admin) & set(rrhh))
    data = {}
    for (municipio, anio) in keys:
        a = admin[(municipio, anio)]
        rh = rrhh.get((municipio, anio), {})
        e = edu.get((municipio, anio), {})
        s = salud.get((municipio, anio), {})

        def g(d, k):
            return d.get(k) or 0

        municipal_total = g(rh, "IRH05") + g(rh, "IRH12") + g(rh, "IRH15") + g(rh, "IRH16")
        educacion_total = g(e, "IEDU040") + g(e, "IEDU042") + g(e, "IEDU043") + g(e, "IEDU041")
        salud_total = g(s, "MPSP") + g(s, "MPSCC") + g(s, "MPSH") + g(s, "MPSCDT") + g(s, "MPSOC")

        gasto_planta = g(a, "IADM78") + g(e, "IEDU040.1") + g(s, "ISAL029")
        gasto_contrata = g(a, "IADM79") + g(e, "IEDU042.1") + g(s, "ISAL031")
        gasto_honorarios = g(a, "IADM80") + g(e, "IEDU043.1") + g(s, "ISAL032")
        gasto_comunitarios = g(a, "IADM111")
        gasto_total = g(a, "IADM61") + g(e, "IEDU026.1") + g(s, "ISAL019.1")

        iadm78 = a.get("IADM78")
        iadm42 = a.get("IADM42")
        lim42 = round(g(a, "IADM61") / (iadm42 * 0.42) * 100, 2) if iadm42 else None
        lim40 = round(g(a, "IADM79") / (iadm78 * 0.40) * 100, 2) if iadm78 else None
        # lim10 = gasto en honorarios / gasto en personal de planta, en % —
        # se compara directo contra el tope legal de 10% (Ley N° 19.280,
        # Art. 13), NO contra 100. Verificado exacto contra Providencia 2024:
        # 1.747.877 / 22.255.103 × 100 = 7,9%.
        # (Antes se usaba IADM80/(IADM61×0,10)×100 — daba 46,46% para el
        # mismo caso, muy por sobre el 10% legal, porque esa fórmula NO es
        # la del Art. 13: dividía por el gasto en personal TOTAL, no por el
        # de planta, y escalaba el resultado contra el 10% en vez de
        # comparar el % real. Se corrige a la fórmula de la ley.)
        lim10 = round(g(a, "IADM80") / iadm78 * 100, 2) if iadm78 else None

        # "Blindspot" de Programas Comunitarios (solo Personal Municipal, no
        # es un límite legal — no tiene techo). El único límite que existe
        # para honorarios es el 10% de la Ley N° 19.280 Art. 13, calculado
        # solo con IADM80/IADM78 (sin Comunitarios). Los Programas
        # Comunitarios (IADM111) están excluidos de todos los límites
        # (dictamen CGR 2017) — sumarlos al numerador muestra cuánto sería
        # el gasto en honorarios si ese vacío legal no existiera.
        # Verificado exacto contra Providencia 2024: (1.747.877 + 3.459.818)
        # / 22.255.103 × 100 = 23,4%.
        blindspot_comunitarios_pct = round((g(a, "IADM80") + g(a, "IADM111")) / iadm78 * 100, 2) if iadm78 else None

        irh15_raw = rh.get("IRH15")
        sindato = (irh15_raw is None or irh15_raw == 0) and g(a, "IADM80") > 0

        gasto_edu_total_sector = e.get("IEDU026.1")
        gasto_sal_total_sector = s.get("ISAL019.1")

        def pct_norm(v):
            """IADM25/IADM33 vienen en escala 0-1 (fracción) en 2008-2023 y
            en escala 0-100 (porcentaje) desde 2024 en adelante — cambio de
            formato de SINIM verificado en las 6.210 filas (0 excepciones a
            este corte por año). Se normaliza siempre a 0-100."""
            if v is None:
                return None
            return round(v * 100, 2) if v <= 1 else round(v, 2)

        data.setdefault(municipio, {})[anio] = {
            "poblacion": poblacion.get((municipio, anio)),
            "profesionalizacion_pct": pct_norm(rh.get("IADM25")),
            "participacion_femenina_pct": pct_norm(rh.get("IADM33")),
            "municipal_total": municipal_total,
            "educacion_total": educacion_total,
            "salud_total": salud_total,
            "consolidado_total": municipal_total + educacion_total + salud_total,
            "consolidado": {
                "planta": g(rh, "IRH05") + g(e, "IEDU040") + g(s, "MPSP"),
                "contrata": g(rh, "IRH12") + g(e, "IEDU042") + g(s, "MPSCC"),
                "cdt": g(e, "IEDU041") + g(s, "MPSCDT"),
                "honorarios": g(rh, "IRH15") + g(e, "IEDU043") + g(s, "MPSH"),
                "comunitarios": g(rh, "IRH16") + g(s, "MPSOC"),
            },
            "municipal_honorarios_sindato": sindato,
            "gasto": {
                "planta": gasto_planta,
                "contrata": gasto_contrata,
                "honorarios": gasto_honorarios,
                "comunitarios": gasto_comunitarios,
                "total": gasto_total,
            },
            "limites": {"lim42": lim42, "lim40": lim40, "lim10": lim10,
                        "blindspot_comunitarios_pct": blindspot_comunitarios_pct},
            "gasto_por_area": {
                "municipal": a.get("IADM61"),
                "educacion": gasto_edu_total_sector,
                "salud": gasto_sal_total_sector,
            },
            # "Activa" = hay gasto en personal del sector O headcount del
            # sector (no solo gasto): hay casos reales (ej. Queilén 2023)
            # donde el sector tiene personas contratadas (IEDU04x/MPSxx > 0)
            # pero el gasto en personal de ese sector viene en 0 en la hoja
            # de Administración/Salud — antes eso hacía desaparecer 246
            # personas reales de la tarjeta de Dotación.
            "areas_activas": {
                "municipal": True,
                "educacion": bool(gasto_edu_total_sector) or educacion_total > 0,
                "salud": bool(gasto_sal_total_sector) or salud_total > 0,
            },
        }

    with open(OUT_JS, "w", encoding="utf-8") as f:
        f.write("// Generado por scripts/build_dotacion.py — no editar a mano.\n")
        f.write("const DATA_DOTACION = ")
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")

    n_comunas = len(data)
    n_rows = sum(len(v) for v in data.values())
    print(f"OK: {n_comunas} comunas, {n_rows} filas comuna-año -> {OUT_JS}")


if __name__ == "__main__":
    main()
