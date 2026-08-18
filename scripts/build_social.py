#!/usr/bin/env python3
"""
Genera data/data_social.js a partir de datos reales SINIM, para las 345
comunas de Chile, años 2008-2025.

Fuentes:
  - /Users/cristobal/prueba/sinim/5-Social_y_comunitaria.xlsx      (CASEN, RSH, organizaciones)
  - /Users/cristobal/prueba/sinim/1-Administracion_finanzas.xlsx   (gasto social: BGMAPSOC, IADM87, IADM88)
  - /Users/cristobal/prueba/sinim/7-caracterizacion comunal.xlsx   (población, ICAR004)

Fórmulas (verificadas exactas contra Providencia 2008/2021):
  casen_pct              = ISOC001
  asistencia_directa     = IADM88
  asistencia_directa_hab = asistencia_directa / poblacion
  asistencia_rm_avg      = promedio de asistencia_directa_hab entre las comunas
                            de la MISMA región, ese año (excluyendo nulos)
  rshnp                  = RSHNP
  rsh60                  = RSHPMA60
  rsh60_pct              = RSHPMA60 / RSHNP * 100
  vulnerabilidad_pct     = RSHNH40 / RSHNHENC * 100
  hogares.total          = RSHNHENC
  hogares.vulnerables    = RSHNH40
  hogares.medios         = RSHNH50 + RSHNH60 + RSHNH70
  hogares.medios_altos   = RSHNH80 + RSHNH90 + RSHNH100
  org_comunitarias       = IADM87
  gasto_prog_sociales    = BGMAPSOC
  gasto_social_total     = (BGMAPSOC o 0) + IADM87
"""
import json
import re
import openpyxl

from build_administracion import comuna_key, num

SINIM_SOCIAL = "/Users/cristobal/prueba/sinim/5-Social_y_comunitaria.xlsx"
SINIM_ADMIN = "/Users/cristobal/prueba/sinim/1-Administracion_finanzas.xlsx"
SINIM_CARAC = "/Users/cristobal/prueba/sinim/7-caracterizacion comunal.xlsx"
REGIONES_JS = "/Users/cristobal/Downloads/maqueta/data/regiones_comunas.js"
OUT_JS = "/Users/cristobal/Downloads/maqueta/data/data_social.js"

SOCIAL_CODES = ["ISOC001", "RSHNH40", "RSHNH50", "RSHNH60", "RSHNH70", "RSHNH80",
                 "RSHNH90", "RSHNH100", "RSHNHENC", "RSHNP", "RSHPMA60"]
ADMIN_CODES = ["BGMAPSOC", "IADM87", "IADM88"]


def find_col(header, code):
    pat = r"(?<![A-Za-z0-9])" + re.escape(code) + r"(?!\.\d)(?![A-Za-z0-9])"
    for i, h in enumerate(header):
        if h and re.search(pat, str(h)):
            return i
    raise KeyError(f"Column not found for code {code}")


def find_anio_col(header):
    for i, h in enumerate(header):
        if h and str(h).strip().upper() == "AÑO":
            return i
    raise KeyError("Año column not found")


def load_sheet(path, codes):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {c: find_col(header, c) for c in codes}
    idx_anio = find_anio_col(header)
    out = {}
    for r in rows:
        if r[0] is None:
            continue
        municipio = comuna_key(str(r[1]).strip())
        anio = str(r[idx_anio])
        out[(municipio, anio)] = {c: num(r[i]) for c, i in idx.items()}
    return out


def load_poblacion():
    wb = openpyxl.load_workbook(SINIM_CARAC, read_only=True)
    ws = wb["Hoja1"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx_pob = [i for i, h in enumerate(header) if h and str(h).startswith("ICAR004")][0]
    idx_anio = len(header) - 1
    out = {}
    for r in rows:
        if r[0] is None:
            continue
        municipio = comuna_key(str(r[1]).strip())
        anio = str(r[idx_anio])
        out[(municipio, anio)] = num(r[idx_pob])
    return out


def load_region_por_comuna():
    with open(REGIONES_JS, encoding="utf-8") as f:
        content = f.read()
    m = re.search(r"const REGION_POR_COMUNA = (\{.*?\});", content, re.S)
    return json.loads(m.group(1))


def main():
    social = load_sheet(SINIM_SOCIAL, SOCIAL_CODES)
    admin = load_sheet(SINIM_ADMIN, ADMIN_CODES)
    poblacion = load_poblacion()
    region_por_comuna = load_region_por_comuna()

    keys = sorted(set(social) & set(admin))
    data = {}
    asistencia_hab_por_comuna_anio = {}

    for key in keys:
        municipio, anio = key
        s = social[key]
        a = admin[key]
        pob = poblacion.get(key)

        asistencia_directa = a.get("IADM88")
        asistencia_directa_hab = (round(asistencia_directa / pob, 3)
                                   if asistencia_directa is not None and pob else None)
        asistencia_hab_por_comuna_anio[key] = asistencia_directa_hab

        rshnp = s.get("RSHNP")
        rsh60 = s.get("RSHPMA60")
        rsh60_pct = round(rsh60 / rshnp * 100, 3) if rsh60 is not None and rshnp else None

        rshnhenc = s.get("RSHNHENC")
        rshnh40 = s.get("RSHNH40")
        vulnerabilidad_pct = (round(rshnh40 / rshnhenc * 100, 3)
                               if rshnh40 is not None and rshnhenc else None)

        def g(k):
            return s.get(k) or 0

        # Solo "total" (RSHNHENC) se deja en None cuando no hay dato (pre-2020,
        # el RSH no existía); las categorías individuales quedan en 0, igual
        # que en los datos originales.
        hogares_vulnerables = g("RSHNH40")
        hogares_medios = g("RSHNH50") + g("RSHNH60") + g("RSHNH70")
        hogares_medios_altos = g("RSHNH80") + g("RSHNH90") + g("RSHNH100")

        gasto_prog_sociales = a.get("BGMAPSOC")
        org_comunitarias = a.get("IADM87")
        gasto_social_total = (gasto_prog_sociales or 0) + (org_comunitarias or 0)
        if gasto_prog_sociales is None and org_comunitarias is None:
            gasto_social_total = None

        data.setdefault(municipio, {})[anio] = {
            "poblacion": pob,
            "casen_pct": s.get("ISOC001"),
            "rshnp": rshnp,
            "rsh60": rsh60,
            "rsh60_pct": rsh60_pct,
            "asistencia_directa": asistencia_directa,
            "asistencia_directa_hab": asistencia_directa_hab,
            "asistencia_rm_avg": None,  # se completa en la segunda pasada
            "vulnerabilidad_pct": vulnerabilidad_pct,
            "hogares": {
                "total": rshnhenc,
                "vulnerables": hogares_vulnerables,
                "medios": hogares_medios,
                "medios_altos": hogares_medios_altos,
            },
            "org_comunitarias": org_comunitarias,
            "gasto_prog_sociales": gasto_prog_sociales,
            "gasto_social_total": gasto_social_total,
        }

    # Segunda pasada: promedio regional de asistencia_directa_hab por año.
    by_region_year = {}
    for (municipio, anio), val in asistencia_hab_por_comuna_anio.items():
        if val is None:
            continue
        region = region_por_comuna.get(municipio)
        if not region:
            continue
        by_region_year.setdefault((region, anio), []).append(val)

    region_avg = {k: round(sum(v) / len(v), 3) for k, v in by_region_year.items()}

    for municipio, years in data.items():
        region = region_por_comuna.get(municipio)
        for anio, rec in years.items():
            if region:
                rec["asistencia_rm_avg"] = region_avg.get((region, anio))

    with open(OUT_JS, "w", encoding="utf-8") as f:
        f.write("// Generado por scripts/build_social.py — no editar a mano.\n")
        f.write("const DATA_SOCIAL = ")
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")

    n_comunas = len(data)
    n_rows = sum(len(v) for v in data.values())
    print(f"OK: {n_comunas} comunas, {n_rows} filas comuna-año -> {OUT_JS}")


if __name__ == "__main__":
    main()
