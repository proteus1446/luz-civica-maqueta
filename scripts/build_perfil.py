#!/usr/bin/env python3
"""
Genera data/data_perfil.js a partir de datos reales SINIM + Contraloría, para
las 345 comunas de Chile, años 2008-2025.

Fuentes:
  - /Users/cristobal/prueba/sinim/7-caracterizacion comunal.xlsx    (población, densidad, % rural)
  - /Users/cristobal/prueba/sinim/5-Social_y_comunitaria.xlsx       (seguridad: cámaras, vehículos, consejo)
  - /Users/cristobal/prueba/sinim/6-Desarrollo_gestion_territorial.xlsx (áreas verdes, vivienda, cultura)
  - /Users/cristobal/prueba/sinim/1-Administracion_finanzas.xlsx    (BGMAPCUL, IADM11 para % cultura)
  - /Users/cristobal/Desktop/luz _c/contraloria .xlsx               (gasto vigilancia)

Fórmulas (verificadas exactas contra Providencia 2008/2020, salvo la excepción
de cultura.gasto anotada abajo):
  densidad                    = ICAR007
  poblacion                   = ICAR004
  rural_pct                   = ICAR008
  seguridad.gasto_vigilancia  = "Servicios de Vigilancia" (Contraloría)
  seguridad.camaras           = MNSEGCIUCA
  seguridad.vehiculos         = MNSEGCIUAU + MNSEGCIUCM + MNSEGCIUMO + MNSEGCIUBI
  seguridad.consejo           = MNSEGCIUCC (texto S-N / No Recepcionado)
  cultura.admin                = CULTADM
  cultura.personal             = CULTPERSO
  cultura.gasto                = BGMAPCUL
  cultura.gasto_pct_total      = BGMAPCUL / IADM11 (gastos.total municipal) * 100
  areas_verdes.m2_hab          = ITER009
  areas_verdes.parques         = MPQC
  areas_verdes.plazas          = MPZC
  areas_verdes.gasto_jardines  = ITER008
  vivienda.agua_conexion       = IVAP
  vivienda.viviendas_censo     = IVC
  vivienda.agua_pct            = agua_conexion / viviendas_censo * 100
  vivienda.permisos            = MNPDEE
  vivienda.recepcion_def       = MCRD
  vivienda.avaluo              = ITER012

NOTA cultura.gasto: el valor original hardcodeado de Providencia 2020
(1.856.425) no coincide con el BGMAPCUL actual del archivo SINIM (1.454.404);
no se encontró ese número en ningún otro archivo disponible (Contraloría,
demo.xlsx). Es muy probable que sea una revisión posterior de SINIM (el
archivo de Administración fue descargado el 13-jul-2025, más reciente que la
maqueta) — se usa el valor SINIM actual, más confiable que el de la maqueta.
"""
import json
import re
import openpyxl

from build_administracion import comuna_key, num

SINIM_CARAC = "/Users/cristobal/prueba/sinim/7-caracterizacion comunal.xlsx"
SINIM_SOCIAL = "/Users/cristobal/prueba/sinim/5-Social_y_comunitaria.xlsx"
SINIM_TERR = "/Users/cristobal/prueba/sinim/6-Desarrollo_gestion_territorial.xlsx"
SINIM_ADMIN = "/Users/cristobal/prueba/sinim/1-Administracion_finanzas.xlsx"
CONTRALORIA = "/Users/cristobal/Desktop/luz _c/contraloria .xlsx"
OUT_JS = "/Users/cristobal/Downloads/maqueta/data/data_perfil.js"


def find_col(header, code):
    pat = r"(?<![A-Za-z0-9])" + re.escape(code) + r"(?!\.\d)(?![A-Za-z0-9])"
    for i, h in enumerate(header):
        if h and re.search(pat, str(h)):
            return i
    raise KeyError(f"Column not found for code {code}")


def find_anio_col(header):
    for i, h in enumerate(header):
        if h and str(h).strip().upper() == "AÑO":
            return i
    raise KeyError("Año column not found")


def load_sheet(path, codes, text_codes=()):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {c: find_col(header, c) for c in codes}
    idx_text = {c: find_col(header, c) for c in text_codes}
    idx_anio = find_anio_col(header)
    out = {}
    for r in rows:
        if r[0] is None:
            continue
        municipio = comuna_key(str(r[1]).strip())
        anio = str(r[idx_anio])
        rec = {c: num(r[i]) for c, i in idx.items()}
        for c, i in idx_text.items():
            v = r[i]
            rec[c] = v if isinstance(v, str) else None
        out[(municipio, anio)] = rec
    return out


def load_contraloria_vigilancia():
    wb = openpyxl.load_workbook(CONTRALORIA, read_only=True)
    ws = wb["Sheet1"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {h: i for i, h in enumerate(header)}
    out = {}
    for r in rows:
        if r[0] is None:
            continue
        anio = str(r[idx["Año"]])
        municipio = comuna_key(str(r[idx["Nombre Municipio"]]).strip())
        out[(municipio, anio)] = num(r[idx["Servicios de Vigilancia"]])
    return out


def main():
    carac = load_sheet(SINIM_CARAC, ["ICAR004", "ICAR007", "ICAR008"])
    social = load_sheet(SINIM_SOCIAL,
                         ["MNSEGCIUCA", "MNSEGCIUAU", "MNSEGCIUCM", "MNSEGCIUMO", "MNSEGCIUBI"],
                         text_codes=["MNSEGCIUCC"])
    terr = load_sheet(SINIM_TERR,
                       ["ITER009", "MPQC", "MPZC", "ITER008", "IVAP", "IVC", "MNPDEE",
                        "MCRD", "ITER012", "CULTPERSO"],
                       text_codes=["CULTADM"])
    admin = load_sheet(SINIM_ADMIN, ["BGMAPCUL", "IADM11"])
    vigilancia = load_contraloria_vigilancia()

    keys = sorted(set(carac) & set(terr))
    data = {}
    for key in keys:
        municipio, anio = key
        c = carac[key]
        t = terr[key]
        s = social.get(key, {})
        a = admin.get(key, {})

        agua_conexion = t.get("IVAP")
        viviendas_censo = t.get("IVC")
        agua_pct = (round(agua_conexion / viviendas_censo * 100, 3)
                    if agua_conexion is not None and viviendas_censo else None)

        gasto_cultura = a.get("BGMAPCUL")
        gasto_muni_total = a.get("IADM11")
        gasto_pct_total = (round(gasto_cultura / gasto_muni_total * 100, 3)
                            if gasto_cultura is not None and gasto_muni_total else None)

        vehiculos = ((s.get("MNSEGCIUAU") or 0) + (s.get("MNSEGCIUCM") or 0) +
                     (s.get("MNSEGCIUMO") or 0) + (s.get("MNSEGCIUBI") or 0))

        data.setdefault(municipio, {})[anio] = {
            "densidad": c.get("ICAR007"),
            "poblacion": c.get("ICAR004"),
            "rural_pct": c.get("ICAR008"),
            "seguridad": {
                "gasto_vigilancia": vigilancia.get(key),
                "camaras": s.get("MNSEGCIUCA"),
                "vehiculos": vehiculos,
                "consejo": s.get("MNSEGCIUCC"),
            },
            "cultura": {
                "admin": t.get("CULTADM"),
                "personal": t.get("CULTPERSO"),
                "gasto": gasto_cultura,
                "gasto_pct_total": gasto_pct_total,
            },
            "areas_verdes": {
                "m2_hab": t.get("ITER009"),
                "parques": t.get("MPQC"),
                "plazas": t.get("MPZC"),
                "gasto_jardines": t.get("ITER008"),
            },
            "vivienda": {
                "agua_conexion": agua_conexion,
                "viviendas_censo": viviendas_censo,
                "agua_pct": agua_pct,
                "permisos": t.get("MNPDEE"),
                "recepcion_def": t.get("MCRD"),
                "avaluo": t.get("ITER012"),
            },
        }

    with open(OUT_JS, "w", encoding="utf-8") as f:
        f.write("// Generado por scripts/build_perfil.py — no editar a mano.\n")
        f.write("const DATA_PERFIL = ")
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")

    n_comunas = len(data)
    n_rows = sum(len(v) for v in data.values())
    print(f"OK: {n_comunas} comunas, {n_rows} filas comuna-año -> {OUT_JS}")


if __name__ == "__main__":
    main()
