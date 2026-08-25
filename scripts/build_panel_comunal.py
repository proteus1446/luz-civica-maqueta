#!/usr/bin/env python3
"""
Panel Comunal (panel_comunal.html) — combina los datos reales ya construidos
para los 6 paneles de detalle (Administración, Dotación, Educación, Salud,
Perfil, Social) en un solo resumen por comuna/año, más 2 fuentes nuevas:
  - Cementerio (9-Cementerio.xlsx): MCEM
  - Sueldo de alcalde (Panel_alcalde_remuneracion(1).xlsx + MGRADALC de
    2-Recursos H.xlsx)

Fórmulas verificadas exactas contra Providencia/Ñuñoa/Maipú del DATA
original hardcodeado de panel_comunal.html:
  administracion.gasto_hab = (gastos.total - (fcm+salud+educacion)) / poblacion
  administracion.delta_pct = (deficit[año] - deficit[año-1]) / |deficit[año-1]| * 100
  administracion.dependencia_fcm    = kpis.dependencia_fcm (de data_administracion.js, = IADM75)
  administracion.deuda_flotante_pagado_pct = kpis.deuda_flotante_pagado_pct (de data_administracion.js)
  educacion.activa   = activa (de data_educacion.js, = bool(IEDU025))
  educacion.administra = administra (de data_educacion.js, viene de MTASE —
    más confiable que activa para decidir "no administra" en la UI, ver
    nota en build_educacion.py)
  salud.administra   = MASM == "Si" (de data_salud.js)
  salud.admin_tipo   = MTAS (de data_salud.js) — a diferencia de Educación,
    en Salud "Corporación" SIEMPRE viene con MASM="Si" (nunca es motivo de
    "no administra": es solo el detalle de quién lo administra dentro de
    los que sí administran). Valores posibles: "Depto. o Dirección",
    "Corporación", "Sin Servicio" (cuando no administra).
  educacion.deficit  = ingresos.total - gastos.total
  educacion.delta_pct = misma fórmula que arriba, sobre educacion.deficit
  educacion.admin_tipo = MTASE (3-Educacion.xlsx, no incluido en build_educacion.py)
  salud.deficit = ISAL009 - ISAL018 (ingresos - gastos totales del sector)
  salud.delta_pct = misma fórmula
  salud.medicos_1000 = MTFCM / HPISM * 1000
  dotacion.planta_pct = consolidado.planta / consolidado_total * 100
  dotacion.func_1000_hab = dotacion.municipal / población × 1.000 (mismo
    criterio que el KPI "Funcionarios por 1.000 habitantes" de
    maqueta_dotacion.html — solo cuenta el área Municipal)
  dotacion.profesionalizacion_pct = profesionalizacion_pct (de
    data_dotacion.js, ya normalizado a escala 0-100)
  dotacion.blindspot_comunitarios_pct = limites.blindspot_comunitarios_pct
    (de data_dotacion.js)
  dotacion.lim42 = limites.lim42 (de data_dotacion.js) — "Participación de
    Gastos en Personal Respecto del Umbral Legal (42%)", criterio SUBDERE:
    (Gasto en Personal Total / (Ingresos Propios × 42%)) × 100. Fuente:
    Balance de Ejecución Presupuestario (BEP). Base legal: Art. 1 Ley
    18.294, modificado por la letra a) del Art. 65 de la Ley 18.382. Incluye
    la renta del alcalde/alcaldesa (va dentro del Subtítulo 21, no se puede
    aislar). Sobre 100% excede el umbral del 42% del ingreso propio.
  dotacion.municipal/educacion/salud = null si areas_activas.<área> es False
    (la comuna no administra ese sector directamente) en vez de 0 — mismo
    criterio que el reparto por área de maqueta_dotacion.html
  dotacion.total = suma solo de las áreas activas (no consolidado_total crudo)
  alcalde.nombre  = nombre_completo (title case)
  alcalde.grado   = MGRADALC (2-Recursos H.xlsx)
  alcalde.mediana/min/max = mediana/minimo/maximo_remuneracion_BRUTA
    (a pedido — el hardcodeado original de Maipú 2024 coincidía con la
    líquida, pero se cambió a bruta deliberadamente; el Excel fuente trae
    ambas si en algún momento se quiere volver a líquida)
  REMUN_LOOKUP[comuna] = {n_total, n_atipicos} desde data_remuneraciones.js
    (nivel comuna, no por año — igual que el original)
"""
import json
import re
import sys

import openpyxl

sys.path.insert(0, ".")
from build_administracion import comuna_key, num
from organismo_mapping import organismo_a_comuna_key

DATA_DIR = "/Users/cristobal/Downloads/maqueta/data"
SINIM_EDU = "/Users/cristobal/prueba/sinim/3-Educacion.xlsx"
SINIM_RRHH = "/Users/cristobal/prueba/sinim/2-Recursos H.xlsx"
SINIM_CEMENTERIO = "/Users/cristobal/prueba/sinim/9-Cementerio.xlsx"
ALCALDE_XLSX = "/Users/cristobal/Downloads/Panel_alcalde_remuneracion(1).xlsx"
OUT_JS = f"{DATA_DIR}/data_panel_comunal.js"


def load_js_object(path, varname):
    with open(path, encoding="utf-8") as f:
        content = f.read()
    m = re.search(r"const " + varname + r" = (\{.*\});", content, re.S)
    return json.loads(m.group(1))


def find_col(header, code):
    pat = r"(?<![A-Za-z0-9])" + re.escape(code) + r"(?!\.\d)(?![A-Za-z0-9])"
    for i, h in enumerate(header):
        if h and re.search(pat, str(h)):
            return i
    raise KeyError(f"Column not found for code {code}")


def find_anio_col(header):
    for i, h in enumerate(header):
        if h and str(h).strip().upper() in ("AÑO", "ANIO"):
            return i
    raise KeyError("Año column not found")


def load_sheet(path, codes, text_codes=()):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {c: find_col(header, c) for c in codes}
    idx_text = {c: find_col(header, c) for c in text_codes}
    idx_anio = find_anio_col(header)
    out = {}
    for r in rows:
        if r[0] is None:
            continue
        municipio = comuna_key(str(r[1]).strip())
        anio = str(r[idx_anio])
        rec = {c: num(r[i]) for c, i in idx.items()}
        for c, i in idx_text.items():
            v = r[i]
            rec[c] = v.strip() if isinstance(v, str) else None
        out[(municipio, anio)] = rec
    return out


def load_alcalde():
    wb = openpyxl.load_workbook(ALCALDE_XLSX, read_only=True)
    ws = wb["Sheet1"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {h: i for i, h in enumerate(header)}
    out = {}
    for r in rows:
        if r[0] is None:
            continue
        municipio = organismo_a_comuna_key("Municipalidad de " + str(r[idx["Comuna"]]).title())
        anio = str(r[idx["anyo"]])
        out[(municipio, anio)] = {
            "nombre": str(r[idx["nombre_completo"]]).title(),
            "mediana": num(r[idx["mediana_remuneracion_bruta"]]),
            "min": num(r[idx["minimo_remuneracion_bruta"]]),
            "max": num(r[idx["maximo_remuneracion_bruta"]]),
        }
    return out


def delta_pct(cur, prev):
    if cur is None or prev is None or prev == 0:
        return None
    return round((cur - prev) / abs(prev) * 100, 3)


def main():
    admin = load_js_object(f"{DATA_DIR}/data_administracion.js", "DATA_ADMINISTRACION")
    dotacion = load_js_object(f"{DATA_DIR}/data_dotacion.js", "DATA_DOTACION")
    educacion = load_js_object(f"{DATA_DIR}/data_educacion.js", "DATA_EDUCACION")
    salud = load_js_object(f"{DATA_DIR}/data_salud.js", "DATA_SALUD")
    social = load_js_object(f"{DATA_DIR}/data_social.js", "DATA_SOCIAL")
    perfil = load_js_object(f"{DATA_DIR}/data_perfil.js", "DATA_PERFIL")
    remun = load_js_object(f"{DATA_DIR}/data_remuneraciones.js", "COMUNAS_REM")
    with open(f"{DATA_DIR}/nombres_comunas.js", encoding="utf-8") as f:
        nombres = json.loads(re.search(r"const NOMBRES_COMUNAS = (\{.*?\});", f.read(), re.S).group(1))

    mtase = load_sheet(SINIM_EDU, [], text_codes=["MTASE"])
    grado_alcalde = load_sheet(SINIM_RRHH, ["MGRADALC"])
    cementerio = load_sheet(SINIM_CEMENTERIO, [], text_codes=["MCEM"])
    alcalde = load_alcalde()

    remun_lookup = {k: {"n_total": v.get("n_total"), "n_atipicos": v.get("n_atipicos")}
                     for k, v in remun.items()}

    # comunas presentes en TODOS los paneles base (administración es el ancla)
    comunas = sorted(set(admin) & set(dotacion) & set(social) & set(perfil))

    data = {}
    for municipio in comunas:
        anios = sorted(admin[municipio].keys())
        for i, anio in enumerate(anios):
            a = admin[municipio].get(anio, {})
            dot = dotacion.get(municipio, {}).get(anio, {})
            edu = educacion.get(municipio, {}).get(anio, {})
            sal = salud.get(municipio, {}).get(anio, {})
            soc = social.get(municipio, {}).get(anio)
            per = perfil.get(municipio, {}).get(anio, {})
            if soc is None:
                continue

            a_prev = admin[municipio].get(str(int(anio) - 1), {})

            gas = a.get("gastos", {}) or {}
            transf_otras = (gas.get("fcm") or 0) + (gas.get("salud") or 0) + (gas.get("educacion") or 0)
            poblacion = a.get("poblacion")
            gas_total = gas.get("total")
            gasto_hab = (round((gas_total - transf_otras) / poblacion, 3)
                         if poblacion and gas_total is not None else None)

            edu_ing = (edu.get("ingresos") or {}).get("total")
            edu_gas = (edu.get("gastos") or {}).get("total")
            edu_deficit = (edu_ing - edu_gas) if edu_ing is not None and edu_gas is not None else None
            edu_prev = educacion.get(municipio, {}).get(str(int(anio) - 1), {})
            edu_ing_prev = (edu_prev.get("ingresos") or {}).get("total")
            edu_gas_prev = (edu_prev.get("gastos") or {}).get("total")
            edu_deficit_prev = (edu_ing_prev - edu_gas_prev) if edu_ing_prev is not None and edu_gas_prev is not None else None

            sal_ing = sal.get("ISAL009")
            sal_gas = sal.get("ISAL018")
            sal_deficit = (sal_ing - sal_gas) if sal_ing is not None and sal_gas is not None else None
            sal_prev = salud.get(municipio, {}).get(str(int(anio) - 1), {})
            sal_ing_prev = sal_prev.get("ISAL009")
            sal_gas_prev = sal_prev.get("ISAL018")
            sal_deficit_prev = (sal_ing_prev - sal_gas_prev) if sal_ing_prev is not None and sal_gas_prev is not None else None

            mtfcm = sal.get("MTFCM")
            hpism = sal.get("HPISM")
            medicos_1000 = round(mtfcm / hpism * 1000, 3) if mtfcm is not None and hpism else None

            dot_total = dot.get("consolidado_total")
            dot_planta = (dot.get("consolidado") or {}).get("planta")
            planta_pct = round(dot_planta / dot_total * 100, 3) if dot_planta is not None and dot_total else None

            # Solo se cuentan las áreas que la comuna efectivamente administra
            # (mismo criterio que usa el reparto por área en maqueta_dotacion.html:
            # areas_activas.* se basa en si el gasto del sector es > 0). Así una
            # comuna donde Educación/Salud las lleva una Corporación externa
            # muestra "No administra" en vez de "0 (0,0%)".
            dot_aa = dot.get("areas_activas") or {}
            dot_muni = dot.get("municipal_total") if dot_aa.get("municipal") else None
            dot_edu = dot.get("educacion_total") if dot_aa.get("educacion") else None
            dot_sal = dot.get("salud_total") if dot_aa.get("salud") else None
            dot_total_activo = (None if dot_muni is None and dot_edu is None and dot_sal is None
                                else (dot_muni or 0) + (dot_edu or 0) + (dot_sal or 0))

            key = (municipio, anio)
            al = alcalde.get(key)
            if al:
                al = dict(al)
                al["grado"] = grado_alcalde.get(key, {}).get("MGRADALC")
                # NOTA (no corregir): la escala oficial de grado alcalde es 1-6,
                # pero el archivo fuente (2-Recursos H.xlsx, MGRADALC) trae 6
                # filas con grado 7 u 8 — siempre un salto aislado de un solo
                # año para la misma persona/comuna que en el resto de los años
                # tiene grado 5 o 6 (ej. CABO DE HORNOS 2020/2021, COCHRANE
                # 2019, COIHUECO 2020, SAAVEDRA 2021, VILLA ALEGRE 2020;
                # 4 de los 6 casos caen en 2020-2021, años de pandemia).
                # Se deja el valor tal cual viene de la fuente oficial SINIM,
                # sin alterarlo — el asistente de IA está avisado de esto
                # vía el glosario en worker/asistente-worker.js.

            data.setdefault(municipio, {})[anio] = {
                "poblacion": poblacion,
                "perfil": {
                    "densidad": per.get("densidad"),
                    "areas_verdes_hab": (per.get("areas_verdes") or {}).get("m2_hab"),
                    "cementerio": cementerio.get(key, {}).get("MCEM"),
                },
                "administracion": {
                    "deficit": a.get("deficit"),
                    "delta_pct": delta_pct(a.get("deficit"), a_prev.get("deficit")),
                    "gasto_hab": gasto_hab,
                    "dependencia_fcm": (a.get("kpis") or {}).get("dependencia_fcm"),
                    "deuda_flotante_pagado_pct": (a.get("kpis") or {}).get("deuda_flotante_pagado_pct"),
                },
                "educacion": {
                    "activa": bool(edu.get("activa")),
                    "administra": edu.get("administra"),
                    "deficit": edu_deficit,
                    "delta_pct": delta_pct(edu_deficit, edu_deficit_prev),
                    "admin_tipo": edu.get("admin_tipo") or mtase.get(key, {}).get("MTASE"),
                    "cobertura": edu.get("cobertura_pct"),
                    "gasto_alumno_mensual": edu.get("gasto_alumno_mensual"),
                    "alumnos_docente": edu.get("alumnos_por_docente"),
                },
                "salud": {
                    "administra": sal.get("MASM") == "Si",
                    "admin_tipo": (sal.get("MTAS") or "").strip() or None,
                    "deficit": sal_deficit,
                    "delta_pct": delta_pct(sal_deficit, sal_deficit_prev),
                    "medicos_1000": medicos_1000,
                    "inscritos_fonasa": hpism,
                    "gasto_inscrito": sal.get("ISAL23"),
                },
                "dotacion": {
                    "total": dot_total_activo,
                    "municipal": dot_muni,
                    "educacion": dot_edu,
                    "salud": dot_sal,
                    "gasto_personal": (dot.get("gasto") or {}).get("total"),
                    "planta_pct": planta_pct,
                    "lim40": (dot.get("limites") or {}).get("lim40"),
                    "lim42": (dot.get("limites") or {}).get("lim42"),
                    "blindspot_comunitarios_pct": (dot.get("limites") or {}).get("blindspot_comunitarios_pct"),
                    "profesionalizacion_pct": dot.get("profesionalizacion_pct"),
                    "func_1000_hab": (round(dot_muni / poblacion * 1000, 2)
                                      if dot_muni is not None and poblacion else None),
                },
                "social": {
                    "casen_pct": soc.get("casen_pct"),
                    "vulnerabilidad_pct": soc.get("vulnerabilidad_pct"),
                    "hogares": soc.get("hogares"),
                    "asistencia_hab": soc.get("asistencia_directa_hab"),
                    "rshnp": soc.get("rshnp"),
                },
                "alcalde": al,
            }

    with open(OUT_JS, "w", encoding="utf-8") as f:
        f.write("// Generado por scripts/build_panel_comunal.py — no editar a mano.\n")
        f.write("const DATA_PANEL_COMUNAL = ")
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")
        f.write("const REMUN_LOOKUP_PANEL = ")
        json.dump(remun_lookup, f, ensure_ascii=False, separators=(",", ":"))
        f.write(";\n")

    n_comunas = len(data)
    n_rows = sum(len(v) for v in data.values())
    print(f"OK: {n_comunas} comunas, {n_rows} filas comuna-año -> {OUT_JS}")


if __name__ == "__main__":
    main()
